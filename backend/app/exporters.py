"""C16 — Export evidence packets and compare tables.

Formats:
  * md      — evidence packet + a "Источники" section (deterministic, always available)
  * jsonld  — schema.org (ScholarlyArticle / Dataset) + PROV-O provenance
              (prov:wasDerivedFrom -> source documents, prov:generatedAtTime, confidence)
  * pdf     — WeasyPrint (md -> HTML -> PDF); degrades to a printable HTML document
              if WeasyPrint's native libs are unavailable
  * xlsx    — openpyxl workbook for compare tables

Binary formats (pdf, xlsx) are returned base64-encoded with an `encoding` flag so the
JSON API contract stays text-only.
"""
from __future__ import annotations

import base64
import io
import json
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ------------------------------------------------------------------ markdown
def to_markdown(query: str, result: Dict[str, Any]) -> str:
    lines: List[str] = []
    lines.append(f"# {query}\n")
    lines.append(result.get("answer_md") or "_нет ответа_")
    lines.append("")
    conf = result.get("confidence_summary")
    conf_txt = conf.get("overall") if isinstance(conf, dict) else conf
    if conf_txt:
        lines.append(f"**Уверенность:** {conf_txt}\n")

    citations = result.get("citations") or []
    if citations:
        lines.append("## Источники\n")
        for i, c in enumerate(citations, 1):
            title = c.get("title") or c.get("doc_id")
            year = f" ({c['year']})" if c.get("year") else ""
            lines.append(f"{i}. **{title}**{year} — `{c.get('doc_id')}`")
            if c.get("quote"):
                lines.append(f"   > {c['quote']}")
        lines.append("")

    contradictions = result.get("contradictions") or []
    if contradictions:
        lines.append("## Противоречия\n")
        for c in contradictions:
            lines.append(f"- {c.get('a_statement')} ⚔ {c.get('b_statement')}")
        lines.append("")

    experts = result.get("experts") or []
    if experts:
        lines.append("## Эксперты по теме\n")
        for e in experts:
            aff = f" — {e['affiliation']}" if e.get("affiliation") else ""
            lines.append(f"- {e.get('name')}{aff} ({e.get('n_works', 0)} работ)")
        lines.append("")

    gaps = result.get("gaps") or []
    if gaps:
        lines.append("## Пробелы\n")
        for g in gaps:
            text = g.get("description") or g.get("title") if isinstance(g, dict) else g
            lines.append(f"- {text}")
        lines.append("")

    lines.append(f"\n---\n_Сгенерировано sci-tangle · {_now()}_")
    return "\n".join(lines)


# ------------------------------------------------------------------ json-ld
def to_jsonld(query: str, result: Dict[str, Any]) -> str:
    """schema.org + PROV-O document with per-source provenance."""
    citations = result.get("citations") or []
    derived: List[Dict[str, Any]] = []
    for c in citations:
        node: Dict[str, Any] = {
            "@type": "CreativeWork",
            "@id": f"urn:doc:{c.get('doc_id')}",
            "name": c.get("title") or c.get("doc_id"),
            "identifier": c.get("doc_id"),
        }
        if c.get("year"):
            node["datePublished"] = str(c["year"])
        if c.get("quote"):
            node["prov:value"] = c["quote"]
        derived.append(node)

    # Dataset if the query looks like a data/gap/compare request, else ScholarlyArticle
    intent = result.get("intent") or {}
    qtype = intent.get("type") or intent.get("query_type")
    main_type = "Dataset" if qtype in {"aggregate", "gap", "compare"} else "ScholarlyArticle"

    doc = {
        "@context": {
            "@vocab": "https://schema.org/",
            "prov": "http://www.w3.org/ns/prov#",
            "xsd": "http://www.w3.org/2001/XMLSchema#",
        },
        "@type": ["prov:Entity", main_type],
        "name": query,
        "abstract": result.get("answer_md"),
        "prov:generatedAtTime": {
            "@type": "xsd:dateTime", "@value": _now(),
        },
        "prov:wasDerivedFrom": derived,
        "citation": [d["@id"] for d in derived],
        "confidence": (result.get("confidence_summary") or {}).get("overall")
        if isinstance(result.get("confidence_summary"), dict)
        else result.get("confidence_summary"),
        "isBasedOn": [
            {"@type": "Person", "name": e.get("name"),
             "affiliation": e.get("affiliation")}
            for e in (result.get("experts") or [])
        ],
        "wasGeneratedBy": {
            "@type": "prov:Activity",
            "prov:used": "sci-tangle knowledge graph (Neo4j + Elasticsearch)",
            "prov:endedAtTime": _now(),
        },
    }
    return json.dumps(doc, ensure_ascii=False, indent=2)


# ------------------------------------------------------------------ pdf
def to_pdf_bytes(query: str, result: Dict[str, Any]) -> (bytes, str):
    """Render the markdown packet to PDF. Returns (bytes, mimetype-ish label).

    Primary path: WeasyPrint (md -> HTML -> PDF). If WeasyPrint's native rendering
    stack is missing at runtime, fall back to a printable HTML document.
    """
    md = to_markdown(query, result)
    try:
        import markdown as md_lib
        html_body = md_lib.markdown(md, extensions=["tables", "fenced_code"])
    except Exception:
        html_body = "<pre>" + md.replace("<", "&lt;") + "</pre>"

    html_doc = f"""<!doctype html><html><head><meta charset="utf-8">
<style>
  body {{ font-family: 'DejaVu Sans', 'Helvetica', sans-serif; font-size: 12px;
          line-height: 1.4; margin: 2em; color: #111; }}
  h1 {{ font-size: 20px; border-bottom: 2px solid #444; padding-bottom: 4px; }}
  h2 {{ font-size: 15px; color: #234; margin-top: 1.2em; }}
  blockquote {{ color: #555; border-left: 3px solid #bbb; margin: 4px 0; padding-left: 8px; }}
  code {{ background: #f2f2f2; padding: 0 3px; }}
</style></head><body>{html_body}</body></html>"""

    try:
        from weasyprint import HTML
        pdf = HTML(string=html_doc).write_pdf()
        return pdf, "pdf"
    except Exception:
        # graceful fallback: return printable HTML (still openable/printable to PDF)
        return html_doc.encode("utf-8"), "html"


# ------------------------------------------------------------------ xlsx
def compare_to_xlsx(compare: Dict[str, Any]) -> bytes:
    """Serialize a /api/compare response into an .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение"

    tech_a = compare.get("tech_a", "A")
    tech_b = compare.get("tech_b", "B")
    header = ["Параметр", tech_a, tech_b]
    ws.append(header)
    hfill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = hfill

    for row in compare.get("rows", []):
        ws.append([row.get("param"), row.get("tech_a"), row.get("tech_b")])

    for col, width in zip("ABC", (28, 45, 45)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def result_to_xlsx(query: str, result: Dict[str, Any]) -> bytes:
    """Serialize a search evidence packet (sources table) into an .xlsx workbook.

    Used when `xlsx` export is requested with a `search_id` (no compare payload),
    so all four formats work off a search result.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Источники"
    ws.append(["Запрос:", query])
    conf = result.get("confidence_summary")
    ws.append(["Уверенность:", conf.get("overall") if isinstance(conf, dict) else conf])
    ws.append([])
    header = ["#", "doc_id", "Заголовок", "Год", "chunk_id", "Цитата"]
    ws.append(header)
    hrow = ws.max_row
    hfill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[hrow]:
        cell.font = Font(bold=True)
        cell.fill = hfill
    for i, c in enumerate(result.get("citations") or [], 1):
        ws.append([i, c.get("doc_id"), c.get("title"), c.get("year"),
                   c.get("chunk_id"), c.get("quote")])
    for col, width in zip("ABCDEF", (5, 14, 40, 8, 18, 70)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ dispatcher
def export(fmt: str, query: str, result: Dict[str, Any],
           compare: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    fmt = (fmt or "md").lower()
    if fmt == "md":
        return {"format": "md", "filename": "evidence_packet.md",
                "content": to_markdown(query, result), "encoding": "text"}
    if fmt == "jsonld":
        return {"format": "jsonld", "filename": "evidence_packet.jsonld",
                "content": to_jsonld(query, result), "encoding": "text"}
    if fmt == "pdf":
        pdf, kind = to_pdf_bytes(query, result)
        ext = "pdf" if kind == "pdf" else "html"
        return {"format": "pdf", "filename": f"evidence_packet.{ext}",
                "content": base64.b64encode(pdf).decode("ascii"),
                "encoding": "base64" if kind == "pdf" else "base64-html"}
    if fmt == "xlsx":
        if compare is not None:
            data = compare_to_xlsx(compare)
            filename = "compare.xlsx"
        else:
            # xlsx off a search result -> evidence sources table
            data = result_to_xlsx(query, result or {})
            filename = "evidence_sources.xlsx"
        return {"format": "xlsx", "filename": filename,
                "content": base64.b64encode(data).decode("ascii"),
                "encoding": "base64"}
    raise ValueError(f"unsupported format '{fmt}'")
